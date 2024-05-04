import time
import pygame as pg
import openpyxl as xl
from datetime import datetime
import sys


pg.init()
pg.mixer.init()

BG = pg.image.load('assets/BG.jpg')
image_b = pg.image.load('assets/imageb.jpg')
image_a = pg.image.load('assets/imagea.jpg')

chime_sfx = pg.mixer.Sound("assets/chime.mp3")

width, height = 1000, 700

screen = pg.display.set_mode((width, height))
# screen = pg.display.set_mode((width, height), pg.FULLSCREEN)

pg.display.set_caption("Digital Timecard Recorder")
icon = pg.image.load('assets/icon.jpg')
pg.display.set_icon(icon)



class Member:
    def __init__(self, name, rect_loc, button_loc, sheet_name):
        self.name = name
        self.button_rect = pg.Rect(rect_loc)
        self.button_loc = button_loc
        self.button_img = image_a
        self.pressed_img = image_b
        self.sheet_name = sheet_name

    @classmethod
    def clock_in_out_message(cls, name):
        screen.blit(BG, (0, 0))
        if clean_time[:2] <= '11':
            text = font.render(f"Good morning {name}. Clocked In: {clean_time}", True, (255, 255, 255))
            text_rect = text.get_rect(center=(width // 2, 50))
        else:
            text = font.render(f"Goodbye {name}. Clocked Out: {clean_time}", True, (255, 255, 255))
            text_rect = text.get_rect(center=(width // 2, 50))
        screen.blit(text, text_rect)
        pg.display.flip()
        time.sleep(2)

    @classmethod
    def input_time(cls, sheet_name):
        db = xl.load_workbook(f'{current_month}.xlsx')
        sheet = db[sheet_name]
        if clean_time[:2] <= '11':
            time_cell = sheet.cell(current_day + 1, 2)
        else:
            time_cell = sheet.cell(current_day + 1, 3)
        time_cell.value = clean_time[:5]
        date_cell = sheet.cell(current_day + 1, 1)
        date_cell.value = current_date
        for column_cells in sheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = length + 2
        db.save(f'{current_month}.xlsx')

    def button_click(self, mouse_pos):
        if self.button_rect.collidepoint(mouse_pos):
            self.button_img = image_b
            Member.input_time(self.sheet_name)
            chime_sfx.play()
            Member.clock_in_out_message(self.name)

    def flip_all(self):
        self.button_img = image_a


def update_display():
    screen.blit(BG, (0, 0))
    text = font.render(f"{current_date} {clean_time}", True, (255, 255, 255))
    text_rect = text.get_rect(center=(width // 2, 50))
    screen.blit(text, text_rect)


def update_buttons():
    for each in staff_list:
        screen.blit(each.button_img, each.button_loc)
        pg.draw.rect(screen, (0, 0, 255), each.button_rect, 2)  # Draw a red rectangle around each button
        text = font.render(f"{each.name}", True, (255, 255, 255))
        text_rect = text.get_rect(center=(each.button_loc[0] + 75, each.button_loc[1] + 25))
        screen.blit(text, text_rect)
    pg.display.flip()


Button_1 = (100, 100, 150, 50)
Button_2 = (300, 100, 150, 50)
Button_3 = (500, 100, 150, 50)
Button_4 = (700, 100, 150, 50)
Button_5 = (100, 200, 150, 50)
Button_6 = (300, 200, 150, 50)
Button_7 = (500, 200, 150, 50)
Button_8 = (700, 200, 150, 50)
Button_9 = (100, 300, 150, 50)
Button_10 = (300, 300, 150, 50)
Button_11 = (500, 300, 150, 50)
Button_12 = (700, 300, 150, 50)
Button_13 = (100, 400, 150, 50)
Button_14 = (300, 400, 150, 50)
Button_15 = (500, 400, 150, 50)
Button_16 = (700, 400, 150, 50)
Button_17 = (100, 500, 150, 50)
Button_18 = (300, 500, 150, 50)
Button_19 = (500, 500, 150, 50)
Button_20 = (700, 500, 150, 50)

# Adding/removing staff
# instantiate them below, giving them a button
# Add/remove them to/from the staff_list list
# Add/remove them to/from the Excel file
# Make sure there is an Excel file for each month

matthew = Member(name='Mr Tranter', rect_loc=Button_1, button_loc=Button_1, sheet_name='Matthew Tranter')
alex = Member(name='Alex', rect_loc=Button_2, button_loc=Button_2, sheet_name='Alex McKinley')
jamie = Member(name='Jamie', rect_loc=Button_3, button_loc=Button_3, sheet_name='Jamie Yugawa')
tanaka = Member(name='Mr Tanaka', rect_loc=Button_4, button_loc=Button_4, sheet_name='Tanaka Yusei')
shiho = Member(name='Shiho', rect_loc=Button_5, button_loc=Button_5, sheet_name='Shiho Tranter')
duncan = Member(name='Mr Miller', rect_loc=Button_6, button_loc=Button_6, sheet_name='Duncan Miller')
blank = Member(name='Blank', rect_loc=Button_7, button_loc=Button_7, sheet_name='Blank Blank')

staff_list = [
    matthew,
    alex,
    jamie,
    tanaka,
    shiho,
    duncan,
    blank
]

run = True
font = pg.font.SysFont("Arial", 36)

while run:
    current_datetime = datetime.now()
    time_frag = (str(current_datetime).split()[1])
    clean_time = time_frag[:8]
    current_date = current_datetime.date()
    current_month = current_datetime.month
    current_day = current_datetime.day
    update_display()
    update_buttons()
    if clean_time[:5] == '12:00':
        for each in staff_list:
            Member.flip_all(each)
    for event in pg.event.get():
        if event.type == pg.QUIT:
            pg.quit()
            sys.exit()
        if event.type == pg.KEYDOWN:
            if event.key == pg.K_ESCAPE:
                run = False
        if event.type == pg.MOUSEBUTTONDOWN:
            mouse_pos = pg.mouse.get_pos()
            for each in staff_list:
                if each.button_rect.collidepoint(mouse_pos):
                    each.button_click(mouse_pos)
